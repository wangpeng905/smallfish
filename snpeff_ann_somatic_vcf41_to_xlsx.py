#!/usr/bin/python3

import os,sys,re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font, Color,Fill


input_vcf_filename = str(sys.argv[1]).split('.')[0]

font = Font(name = 'Arial',size=10,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')

xlsx_wb = Workbook()
xlsx_wb_filename = input_vcf_filename + '.' + 'xlsx'
xlsx_wb.active.title = input_vcf_filename[:30]
xlsx_wb_sheet = xlsx_wb.active
xlsx_wb_sheet.sheet_properties.tabColor = "FF9966"

greenFill = PatternFill(start_color='CCFF99',
                      end_color='CCFF99',
                      fill_type='solid')
for i in range(1,13):
    xlsx_wb_sheet.cell(row = 1, column = i).fill = greenFill
    xlsx_wb_sheet.cell(row = 1, column = i).font = font

xlsx_wb_sheet.cell(row = 1, column = 1).value = "Chrom"
xlsx_wb_sheet.cell(row = 1, column = 2).value = "Pos"
xlsx_wb_sheet.cell(row = 1, column = 3).value = "Ref_Allele"
xlsx_wb_sheet.cell(row = 1, column = 4).value = "Alt_Allele"
xlsx_wb_sheet.cell(row = 1, column = 5).value = "Somatic Genotype"
xlsx_wb_sheet.cell(row = 1, column = 6).value = "Somatic Variant Quality"
xlsx_wb_sheet.cell(row = 1, column = 7).value = "Gene"
xlsx_wb_sheet.cell(row = 1, column = 8).value = "HGVS(DNA level)"
xlsx_wb_sheet.cell(row = 1, column = 9).value = "HGVS(Protein level)"
xlsx_wb_sheet.cell(row = 1, column = 10).value = "Variant Annotation"
xlsx_wb_sheet.cell(row = 1, column = 11).value = "Allele | Annotation | Annotation_Impact | Gene_Name | Gene_ID | Feature_Type | Feature_ID | Transcript_BioType | Rank | HGVS.c | HGVS.p | cDNA.pos / cDNA.length | CDS.pos / CDS.length | AA.pos / AA.length | Distance | ERRORS / WARNINGS / INFO"
xlsx_wb_sheet.cell(row = 1, column = 12).value = "COSMICv78"

#冻结首行，设置列宽
xlsx_wb_sheet.freeze_panes = 'A2'

for j in ['A','B','C','D','E','F','G','H','I','J','K','L']:
    xlsx_wb_sheet.column_dimensions[j].width = 15


with open(sys.argv[1],'r') as input_vcf:
    rowNum = 2
    
    for line in input_vcf:
        if line.startswith("#"):
            pass

        else:
            CHROM,POS,ID,REF,ALT,QUAL,FILTER,INFO,FORMAT,NORMAL,TUMOR = line.strip().split("\t")
            INFO = INFO.split(';')
            QSS = ''
            SGT = ''
            snpeff_INFO = ''
            Gene = ''
            Variant_region = ''
            HGVS_c = ''
            HGVS_p = ''

            for item in INFO:
                if item.startswith('QSS='):
                    QSS = item.split('=')[-1]
                elif item.startswith('SGT='):
                    SGT = item.split('=')[-1]
                elif item.startswith('ANN='):
                    snpeff_INFO = item
                    Gene = item.split('|')[3]
                    HGVS_c = item.split('|')[9]
                    HGVS_p = item.split('|')[10]

                elif item.startswith('LOF='):
                    Gene = item.split('|')[1]
                elif item.startswith('NMD='):
                    Gene = item.split('|')[1]

            if snpeff_INFO:
                Variant_region = snpeff_INFO.split('|')[1]
            #for item in snpeff_INFO.split(','):
            #    Gene = Gene + item.split('|')[3]
            #    Variant_region = Variant_region + item.split('|')[1]

            xlsx_wb_sheet.cell(row = rowNum, column = 1).value = CHROM
            xlsx_wb_sheet.cell(row = rowNum, column = 2).value = POS
            xlsx_wb_sheet.cell(row = rowNum, column = 3).value = REF
            xlsx_wb_sheet.cell(row = rowNum, column = 4).value = ALT
            xlsx_wb_sheet.cell(row = rowNum, column = 5).value = SGT
            xlsx_wb_sheet.cell(row = rowNum, column = 6).value = QSS
            xlsx_wb_sheet.cell(row = rowNum, column = 7).value = Gene
            xlsx_wb_sheet.cell(row = rowNum, column = 8).value = HGVS_c
            xlsx_wb_sheet.cell(row = rowNum, column = 9).value = HGVS_p
            xlsx_wb_sheet.cell(row = rowNum, column = 10).value = Variant_region
            xlsx_wb_sheet.cell(row = rowNum, column = 11).value = snpeff_INFO
            xlsx_wb_sheet.cell(row = rowNum, column = 12).value = ID
            rowNum += 1

xlsx_wb.save(xlsx_wb_filename)

